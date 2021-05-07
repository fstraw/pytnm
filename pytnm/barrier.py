# -*- coding: utf-8 -*-
"""
Purpose: Quickly assess the feasibility and reasonableness 
of a given barrier design, given a reduction design goal,
a reasonable cost, and any other state-specific criteria
"""

import os
import sys
import openpyxl as p


def _xlsx_to_traffic(ws):
    """Converts .xlsx spreadsheet of TNM Traffic to dictionary
	
	Arguments:
		ws {openpyxl.worksheet} -- [description]
	
	Returns:
		[type] -- [description]
	"""
    filter_rows = 7 ##omit these rows at start of spreadsheet    
    traf_ws = ws
    traf_dict = dict([(r[1].value.strip(), ((r[5].value, r[6].value), (r[7].value, r[8].value), (r[9].value, r[10].value)))
                    for r in traf_ws.rows if r[1].value][filter_rows:])
    return traf_dict

def _barriers_to_list(ws):
    """Returns filtered list of barriers
	
	Arguments:
		ws {openpyxl.worksheet} -- [description]
	
	Returns:
		[type] -- [description]
	"""
    barriers_ws = ws
    bars = []
    for row in barriers_ws.rows:
        vals = [cell.value for cell in row]
        bars.append(vals)
    bars_no_header = bars[15:] # clip TNM2.5 header
    barriers = []
    for bar in bars_no_header:
        bar_name, bar_max_height, square_foot_cost = bar[1], bar[4], bar[5]
        x, y, z = bar[13], bar[14], bar[15]
        # check for empty end rows
        if None in (x, y, z):
            break
        if bar_name:
            barriers.append((bar_name, bar_max_height, square_foot_cost, [(x, y, z)]))
            found_bar = True
        else:
            found_bar = False
            if not found_bar:
                barriers[-1][3].append((x, y, z))
    return barriers

def _receivers_to_list(ws):
	receivers_raw = []	
	for row in ws.rows:
		vals = [cell.value for cell in row]
		receivers_raw.append(vals)
	receivers_no_header = receivers_raw[16:] # clip TNM2.5 header
	receivers = []
	for receiver in receivers_no_header:
		rec_id = receiver[1]
		receptors = receiver[3]
		x, y, z = receiver[4], receiver[5], receiver[6]
		height = receiver[7]
		nac_level = receiver[9]
		if None in (x, y, z):
			break
		else:
			receivers.append((rec_id.strip(), receptors, height, nac_level, (x, y, z)))
	return receivers
			
def _rds_to_list(ws):
    """
    Returns filtered list of roads    
    :param ws: 
    :return: road_list
    """    
    roads_ws = ws
    rds = []
    for row in roads_ws.rows:
        vals = [cell.value for cell in row]
        rds.append(vals)
    rds_no_header = rds[15:] # clip TNM2.5 header
    roads = []
    for rd in rds_no_header:
        rd_name = rd[1]
        rd_width = rd[2]
        pnt = rd[4].strip()
        x, y, z = rd[6], rd[7], rd[8]
        # check for empty end rows
        if None in (x, y, z):
            break
        if rd_name:
            roads.append([rd_name.strip(), rd_width, [[pnt,[x, y, z]]]])
            found_road = True
        else:
            found_road = False
            if not found_road:
                roads[-1][2].append([pnt,[x, y, z]])
    return roads

def _validate_xlsx(xlsx):
	pass

def _read_roadways(xlsx):
	workbook = p.load_workbook(xlsx)
	try:
		roadways_worksheet = [ws for ws in workbook.worksheets if ws['B5'].value == 'INPUT: ROADWAYS'][0]
	except IndexError:
		print('Roadways worksheet not found')
	return(_rds_to_list(roadways_worksheet))

def _read_traffic(xlsx):
	workbook = p.load_workbook(xlsx)
	try:
		traffic_worksheet = [ws for ws in workbook.worksheets if ws['B5'].value == 'INPUT: TRAFFIC FOR LAeq1h Volumes'][0]
	except IndexError:
		print('Traffic worksheet not found')
	return(_xlsx_to_traffic(traffic_worksheet))

def _read_receivers(xlsx):
	workbook = p.load_workbook(xlsx)
	try:
		receiver_worksheet = [ws for ws in workbook.worksheets if ws['B5'].value == 'INPUT: RECEIVERS'][0]
	except IndexError:
		print('Receiver worksheet not found')
	return(_receivers_to_list(receiver_worksheet))

def _read_barriers(xlsx):
	workbook = p.load_workbook(xlsx)
	try:
		barrier_worksheet = [ws for ws in workbook.worksheets if ws['B5'].value == 'INPUT: BARRIERS'][0]
	except IndexError:
		print('Barrier worksheet not found')
	return(_barriers_to_list(barrier_worksheet))

class VehicleClassification(object):
	def __init__(self, volume, speed):
		self.volume = volume
		self.speed = speed

class Auto(VehicleClassification):
	def __repr__(self):
		return 'Auto'

class Medium(VehicleClassification):
	def __repr__(self):
		return 'Medium'

class Heavy(VehicleClassification):
	def __repr__(self):
		return 'Heavy'

class Traffic(object):
	"""Represents Traffic TNM object
	
	Arguments:
		object {[type]} -- [description]
	
	Returns:
		[type] -- [description]
	"""
	def __init__(self, auto, medium, heavy, bus=None, motorcycle=None):
		self.auto = auto
		self.medium = medium
		self.heavy = heavy
		self.bus = bus
		self.motorcycle = motorcycle
	def __repr__(self):
		return self.auto

class Roadway(object):
	"""Represents Roadway TNM object
	
	Arguments:
		object {[type]} -- [description]
	
	Returns:
		[type] -- [description]
	"""
	def __init__(self, name, width, points, traffic=None):
		self.name = name
		self.width = width
		self.points = points
		self.traffic = traffic
		self.geometry = None
	def __repr__(self):
		return self.name


class Receiver(object):
	"""Represents Receiver TNM object
	
	Arguments:
		object {[type]} -- [description]
	"""
	def __init__(self, rec_id, receptors, height, nac_level, geometry, nac=None):
		self.rec_id = rec_id
		self.receptors = receptors
		self.height = height
		self.nac_level = nac_level
		self.geometry = geometry
	def __repr__(self):
		return self.rec_id

class Barrier(object):
	"""Represents Barrier TNM object
	
	Arguments:
		object {[type]} -- [description]
	
	Returns:
		[type] -- [description]
	"""
	def __init__(self, name, max_height, geometry, square_foot_cost=0):
		self.name = name
		self.max_height = max_height
		self.geometry = geometry
		self.square_foot_cost = square_foot_cost
	def __repr__(self):
		return self.name

class TerrainLine(object):
	pass

class BuildingRow(object):
	pass

class Model(object):
	"""[summary]
	
	>> build_run = pytnm.Model(string(.xlsx))
	>> print(build_run.info)
	>> '(('Build Run', 'Project', 'Company', 'Analyst', 'Date'))
	>> roadways = build_run.roadways
	>> roadway = roadways[0]
	>> print(roadway.name)
	>> 'I-10 On Ramp 1'
	>> print(roadway.width)
	>> 22
	>> print(roadway.geometry)
	>> ((123456, 4355422, 12), (123456, 4355422, 15), (123456, 4355422, 12))
	>> print(roadway.traffic)
	>> (125, 10, 5, 3, 1)
	>> print(roadway.traffic.auto)
	>> 125
	>> print(roadway.traffic.auto.speed)
	>> 45

	Arguments:
		object {[type]} -- [description]
	"""

	def __init__(self, xlsx):
		self.TRAFFIC = _read_traffic(xlsx)
		self._roads = _read_roadways(xlsx)
		self._receivers = _read_receivers(xlsx)
		self._barriers = _read_barriers(xlsx)
	
	@property
	def barriers(self):
		barriers = []
		for barrier in self._barriers:
			name = barrier[0]
			max_height = barrier[1]
			geometry = barrier[2]
			barriers.append(Barrier(name, max_height, geometry))		
		return barriers

	@property
	def receivers(self):
		receivers = []		
		for receiver in self._receivers:
			rec_id = receiver[0]
			receptors = receiver[1]			
			height = receiver[2]
			nac_level = receiver[3]
			x, y, z = receiver[4]
			receivers.append(Receiver(rec_id, receptors, height, nac_level, geometry=(x, y, z)))
		return receivers

	@property	
	def roadways(self):
		roads = []
		for road in self._roads:
			name = road[0]
			width = road[1]
			points = road[2]
			autos = self.TRAFFIC[name][0]
			medium = self.TRAFFIC[name][1]
			heavy = self.TRAFFIC[name][2]
			traffic = Traffic(Auto(autos[0], autos[1]), Medium(medium[0], medium[1]), Heavy(heavy[0], heavy[1]))
			roads.append(Roadway(name, width, points, traffic))
		return roads		

class Analysis(object):
	"""Barrier analysis class
	
	This class will contain several read-only properties
	that will aid in documenting the feasibility and 
	reasonableness of a given barrier design modeled in the FHWA
	TNM 2.5. Information is pulled from an ".xlsx" wb of any name 
	with the assumption that the Barrier Design Table results and
	Sound Level Results tables are copied AS-IS directly from TNM 2.5
	to two separate worksheets in the ".xlsx" workbook.
	
	:param wbname: ".xlsx" file on disk (i.g. "Barriers.xlsx")
	:param sndsheet: "worksheet in ".xlsx" containing Sound Level Results
	:param barriercost: cost of barrier design, as reported by TNM 2.5
	"""

	def __init__(self, wbname, sndsheet, barcost=0):
		self._wbname = wbname
		self._sndsheet = sndsheet
		self.barriercost = barcost		
		self._sndrecs = self._wbhandler(sndsheet)
		"""Receivers listed in the Sound Results worksheet"""
		self.snd_rec_list = [(f"{i[0].value}".strip(), i[2].value, i[8].value.strip()) for i in self._sndrecs]
		self.impacted_recs = [i for i in self.snd_rec_list if i[2] != "----"]
		self.impact_num = sum([tup[1] for tup in self.impacted_recs])
		self.du_analysis = sum([tup[1] for tup in self.recs_analysis])
		self.benefitted =  [
                                  (i[0], i[1]) for i in self.recs_analysis 
                                      if i[2] >= 5
                              ]
		self.benefit_num = sum([tup[1] for tup in self.benefitted])
		self.ben_and_imp = [
                                  (i[0], i[1]) for i in self.recs_analysis 
                                      if i[2] >= 5 and i[4] != " ----"
                               ]
		self.ben_and_imp_num = sum([tup[1] for tup in self.ben_and_imp])
		self.reas_red_recs = [
                                  (i[0], i[1]) for i in self.recs_analysis
                                      if i[2] >= i[3]
                                ]
		self.reas_red_num = sum([tup[1] for tup in self.reas_red_recs])
	def _wbhandler(self, sht):
		"""
		Load Excel workbook and generate appropriate receiver list
		depending on the type of results contained in the worksheet
		"""
		wb = p.load_workbook(self._wbname)
		ws = wb.get_sheet_by_name(sht)
		if not ws:
			raise ValueError("Excel worksheet not found!" + 
                                       "Is it spelled correctly?")
		# omit last 8 rows in snd results
		lastrow = ws.max_row - 8
		# limit to appropriate columns 
		reclist = list(ws.iter_rows(min_col=2, max_col=14, min_row=20, max_row=lastrow))
		return reclist
	@property		
	def recs_analysis(self):
		"""
		List receivers, DU, barrier reduction value, reduction goal, and
		impact status in chosen barrier analysis
		"""
		sndreclist = self._sndrecs
		return [("1A", 1, 5.5, 8, " ----")]
		# return (rec, du, barred, redgoal, impstat)
	@property
	def benefits(self):
		"""
		Return tuple of benefited receiver/receptor counts
		"""
		return [(r[0], r[1]) for r in self.recs_analysis
					if r[2] >= 5]
	@property
	def perc_imp_benefitted(self):
		"""
		Return percentage of total receivers in barrier analysis that
		receive a 5 dBA reduction
		"""
		perc = float(self.ben_and_imp_num) / float(self.impact_num)
		return "{:.0%}".format(perc)
	@property
	def perc_ben_reasonable(self):
		"""
		Return percentage of benefits in barrier analysis that
		receive an reasonable reduction
		"""
		if self.benefit_num == 0:
			perc = 0
		else:
			perc = float(self.reas_red_num) / float(self.benefit_num)
		return "{:.0%}".format(perc)
	@property
	def feasible(self):
		"""
		Barrier design is feasible if 75% or more of impacted receivers
		receive a 5dBA or more noise reduction
		"""
		perc_crit = self.impact_num * 0.75
		if self.benefit_num >= perc_crit:
			return True
		else:
			return False
	@property
	def reasonable(self):
		"""
		Barrier design is reasonable if 80% or more of benefitted receivers
		receive a 8dBA or more noise reduction
		"""
		perc_crit = self.benefit_num * 0.80
		if perc_crit == 0:
			return False
		elif self.reas_red_num >= perc_crit:
			return True
		else:
			return False
	def cost_per_benefit(self):
		"""
		If a barrier cost has been specified, returns the cost per 
		benefitted receptor based on the policy-specific allowable cost
		"""
		if self.barriercost <= 0:
		    return 0
		else:
		    return self.barriercost / self.benefit_num


# class GAAnalysis(Analysis):
#     def __init__(self, wbname, barsheet, sndsheet, barcost=0):
#         super(GAAnalysis, self).__init__(wbname, barsheet, sndsheet, barcost=0)
# 	@property
# 	def feasible(self):
# 		"""
#         Barrier design is feasible if one or more of impacted receivers
#         receive a 5dBA or more noise reduction
#         """
#         if self.ben_and_imp_num >= 1:
#             return True
#         return False

# 	@property
# 	def reasonable(self):
# 		"""
#         Barrier design is reasonable if 80% or more of benefitted receivers
#         receive a 8dBA or more noise reduction
#         """
# 		perc_crit = self.benefit_num * 0.80
# 		if perc_crit == 0:
# 			return False
# 		elif self.reas_red_num >= perc_crit:
# 			return True
# 		else:
# 			return False

class LouisianaAnalysis(Analysis):

	def barrier_segments(self):
		"""Read "Barrier_Segments" worksheet and return raw list of rows

		Raises:
			ValueError: Workbook needs to contain a worksheet called "Barrier Segments"

		Returns:
			[type]: [description]
		"""
		wb = p.load_workbook(self._wbname)
		ws = wb.get_sheet_by_name("Barrier_Segments")
		if not ws:
			raise ValueError("Excel worksheet not found!" + 
                                       "Is it spelled correctly?")
		lastrow = ws.max_row
		# limit to appropriate columns 
		barrierlist = list(ws.iter_rows(min_col=2, max_col=14, min_row=16, max_row=lastrow))
		return barrierlist
	
	def _barrier_infos(self):
		"""
		For each barrier in "Barrier_Segments", compiles list of barrier dimensions per segment		
		Returns:
			list: [[{ barrier: [(hgt, length, sqft)] }]]
		"""
		barrier_list = []
		barrier_info = {}
		barrier_tag = ""
		row_count = (len(self.barrier_segments()))
		for idx, row in enumerate(self.barrier_segments()):
			barrier_name = row[0].value
			barrier_point = row[4].value
			barrier_height = row[6].value
			barrier_length = row[8].value
			barrier_sq_footage = row[9].value			
			if (barrier_name and barrier_point): # check for barrier name
				barrier_tag = barrier_name.strip()
				barrier_dimensions = (barrier_length, barrier_height, barrier_sq_footage)
				barrier_info[barrier_name.strip()] = []			
				barrier_info[barrier_name.strip()].append(barrier_dimensions)
			elif barrier_point:
				barrier_dimensions = (barrier_length, barrier_height, barrier_sq_footage)
				if (idx == row_count - 1): # deal with last row					
					barrier_info[barrier_tag].append(barrier_dimensions)
					barrier_list.append(barrier_info)
				else:
					barrier_info[barrier_tag].append(barrier_dimensions)
			elif not barrier_point:				
				barrier_list.append(barrier_info)		
				barrier_info = {}
		return barrier_list

	def _barrier_dimensions(self):
		"""[summary]

		Returns:
			[dict]: [{ barrier: {hgt: sqfootage, hgt:sqfootage, ...} }]
		"""
		barrier_info = {}
		for barrier in self._barrier_infos():
			barrier_name = [k for k in barrier.keys()][0]
			barrier_segment_info = {}
			segment_descriptions_list = barrier.values()
			for segment_descriptions in segment_descriptions_list:
				unique_heights = set([i[1] for i in segment_descriptions if i[1]]) #exclude zero height segments
				for height in unique_heights:
					square_footage = sum([i[2] for i in segment_descriptions if i[1] == height])
					barrier_segment_info[height] = square_footage
			barrier_info[barrier_name] = barrier_segment_info
		return barrier_info

	def barrier_segment_summary(self, barriers):
		"""Given list of barriers, filter barrier dimension dictionary

		Args:
			barriers ([type]): [description]

		Returns:
			[type]: [description]
		"""
		barrier_selection = {}
		for barrier in barriers:
			try:
				barrier_selection[barrier] = self._barrier_dimensions()[barrier]
			except KeyError:
				pass
		return barrier_selection

	def total_square_footage(self, barriers):
		"""Given a list of barriers or single barrier in this analysis, return the total square footage 

		Args:
			barriers (list, string): [description]

		Returns:
			[int]: [description]
		"""
		if type(barriers) == list:
			total_square_footage = 0
			for barrier in barriers:
				barrier_dimensions = self._barrier_dimensions()[barrier]
				total_square_footage += sum([sf for sf in barrier_dimensions.values()])
		else:
			barrier_dimensions = self._barrier_dimensions()[barriers]
			total_square_footage = sum([sf for sf in barrier_dimensions.values()])
		return total_square_footage

	def total_length(self, barriers):
		"""Given list of barriers, return total length

		Args:
			barriers ([type]): [description]

		Returns:
			[type]: [description]
		"""
		barrier_info = {}
		total_length = 0
		for barrier in self._barrier_infos():
			barrier_name = [k for k in barrier.keys()][0]
			if barrier_name in barriers:
				segment_descriptions_list = list(barrier.values())
				total_length += sum([i[0] for i in segment_descriptions_list[0]])
		return total_length

	def min_max_height(self, barriers):
		"""Given a list of barriers, what is the min height

		Args:
			barriers ([type]): [description]

		Returns:
			[type]: [description]
		"""
		barrier_info = {}
		all_heights = []
		for barrier in self._barrier_infos():
			barrier_name = [k for k in barrier.keys()][0]
			if barrier_name in barriers:
				segment_descriptions_list = list(barrier.values())				
				barrier_heights = [i[1] for i in segment_descriptions_list[0] if i[1] !=0] # ignore zero height segments
				all_heights += barrier_heights
		return min(all_heights), max(all_heights)

	@property
	def feasible(self):
		"""
		Barrier design is feasible if 75% or more of impacted receivers
		receive a 5dBA or more noise reduction
		"""
		perc_crit = self.impact_num * 0.75
		if self.benefit_num >= perc_crit:
			return True
		else:
			return False
	@property
	def reasonable(self):
		"""
		Barrier design is reasonable if 80% or more of benefitted receivers
		receive a 8dBA or more noise reduction
		"""
		perc_crit = self.benefit_num * 0.80
		if perc_crit == 0:
			return False
		elif self.reas_red_num >= perc_crit:
			return True
		else:
			return False

def cost_per_square_foot(hgt, total_sqft):
	"""Given total square footage and barrier height, return cost per square foot
		DOTD Cost Table 2016
		This is ugly, but functional

	Args:
		total_sq_foot ([type]): [description]
		height ([type]): [description]
	"""	
	if hgt <= 10:
		if total_sqft <= 10000:
			cost_per_square_foot = 27
		elif 10001 <= total_sqft <= 15000:
			cost_per_square_foot = 26
		elif 15001 <= total_sqft <= 20000:
			cost_per_square_foot = 24
		elif 20001 <= total_sqft <= 25000:
			cost_per_square_foot = 23
		elif 25001 <= total_sqft <= 30000:
			cost_per_square_foot = 22
		elif 30001 <= total_sqft <= 35000:
			cost_per_square_foot = 22
		elif 35001 <= total_sqft <= 40000:
			cost_per_square_foot = 21
		elif 40001 <= total_sqft <= 45000:
			cost_per_square_foot = 21
		elif 45001 <= total_sqft <= 50000:
			cost_per_square_foot = 20
		elif 50001 <= total_sqft <= 55000:
			cost_per_square_foot = 20
		elif 55001 <= total_sqft <= 60000:
			cost_per_square_foot = 20
		elif 60001 <= total_sqft <= 65000:
			cost_per_square_foot = 20
		elif 65001 <= total_sqft <= 70000:
			cost_per_square_foot = 19
		elif 70001 <= total_sqft <= 75000:
			cost_per_square_foot = 19
		elif 75001 <= total_sqft <= 80000:
			cost_per_square_foot = 19
		elif total_sqft >= 80001:
			cost_per_square_foot = 19																																														
	elif 11 <= hgt <= 14:
		if total_sqft <= 10000:
			cost_per_square_foot = 43
		elif 10001 <= total_sqft <= 15000:
			cost_per_square_foot = 41
		elif 15001 <= total_sqft <= 20000:
			cost_per_square_foot = 39
		elif 20001 <= total_sqft <= 25000:
			cost_per_square_foot = 37
		elif 25001 <= total_sqft <= 30000:
			cost_per_square_foot = 36
		elif 30001 <= total_sqft <= 35000:
			cost_per_square_foot = 35
		elif 35001 <= total_sqft <= 40000:
			cost_per_square_foot = 34
		elif 40001 <= total_sqft <= 45000:
			cost_per_square_foot = 33
		elif 45001 <= total_sqft <= 50000:
			cost_per_square_foot = 33
		elif 50001 <= total_sqft <= 55000:
			cost_per_square_foot = 32
		elif 55001 <= total_sqft <= 60000:
			cost_per_square_foot = 32
		elif 60001 <= total_sqft <= 65000:
			cost_per_square_foot = 31
		elif 65001 <= total_sqft <= 70000:
			cost_per_square_foot = 31
		elif 70001 <= total_sqft <= 75000:
			cost_per_square_foot = 30
		elif 75001 <= total_sqft <= 80000:
			cost_per_square_foot = 30
		elif total_sqft >= 80001:
			cost_per_square_foot = 30
	elif 15 <= hgt <= 19:
		if total_sqft <= 10000:
			cost_per_square_foot = 78
		elif 10001 <= total_sqft <= 15000:
			cost_per_square_foot = 76
		elif 15001 <= total_sqft <= 20000:
			cost_per_square_foot = 71
		elif 20001 <= total_sqft <= 25000:
			cost_per_square_foot = 68
		elif 25001 <= total_sqft <= 30000:
			cost_per_square_foot = 66
		elif 30001 <= total_sqft <= 35000:
			cost_per_square_foot = 64
		elif 35001 <= total_sqft <= 40000:
			cost_per_square_foot = 62
		elif 40001 <= total_sqft <= 45000:
			cost_per_square_foot = 61
		elif 45001 <= total_sqft <= 50000:
			cost_per_square_foot = 60
		elif 50001 <= total_sqft <= 55000:
			cost_per_square_foot = 59
		elif 55001 <= total_sqft <= 60000:
			cost_per_square_foot = 58
		elif 60001 <= total_sqft <= 65000:
			cost_per_square_foot = 57
		elif 65001 <= total_sqft <= 70000:
			cost_per_square_foot = 56
		elif 70001 <= total_sqft <= 75000:
			cost_per_square_foot = 56
		elif 75001 <= total_sqft <= 80000:
			cost_per_square_foot = 55
		elif total_sqft >= 80001:
			cost_per_square_foot = 55
	elif 20 <= hgt <= 25:
		if total_sqft <= 10000:
			cost_per_square_foot = 151
		elif 10001 <= total_sqft <= 15000:
			cost_per_square_foot = 145
		elif 15001 <= total_sqft <= 20000:
			cost_per_square_foot = 136
		elif 20001 <= total_sqft <= 25000:
			cost_per_square_foot = 130
		elif 25001 <= total_sqft <= 30000:
			cost_per_square_foot = 126
		elif 30001 <= total_sqft <= 35000:
			cost_per_square_foot = 122
		elif 35001 <= total_sqft <= 40000:
			cost_per_square_foot = 119
		elif 40001 <= total_sqft <= 45000:
			cost_per_square_foot = 117
		elif 45001 <= total_sqft <= 50000:
			cost_per_square_foot = 115
		elif 50001 <= total_sqft <= 55000:
			cost_per_square_foot = 113
		elif 55001 <= total_sqft <= 60000:
			cost_per_square_foot = 111
		elif 60001 <= total_sqft <= 65000:
			cost_per_square_foot = 110
		elif 65001 <= total_sqft <= 70000:
			cost_per_square_foot = 108
		elif 70001 <= total_sqft <= 75000:
			cost_per_square_foot = 107
		elif 75001 <= total_sqft <= 80000:
			cost_per_square_foot = 106
		elif total_sqft >= 80001:
			cost_per_square_foot = 105
	elif hgt >= 26:
		if total_sqft <= 10000:
			cost_per_square_foot = 229
		elif 10001 <= total_sqft <= 15000:
			cost_per_square_foot = 221
		elif 15001 <= total_sqft <= 20000:
			cost_per_square_foot = 208
		elif 20001 <= total_sqft <= 25000:
			cost_per_square_foot = 199
		elif 25001 <= total_sqft <= 30000:
			cost_per_square_foot = 192
		elif 30001 <= total_sqft <= 35000:
			cost_per_square_foot = 186
		elif 35001 <= total_sqft <= 40000:
			cost_per_square_foot = 182
		elif 40001 <= total_sqft <= 45000:
			cost_per_square_foot = 178
		elif 45001 <= total_sqft <= 50000:
			cost_per_square_foot = 175
		elif 50001 <= total_sqft <= 55000:
			cost_per_square_foot = 172
		elif 55001 <= total_sqft <= 60000:
			cost_per_square_foot = 169
		elif 60001 <= total_sqft <= 65000:
			cost_per_square_foot = 167
		elif 65001 <= total_sqft <= 70000:
			cost_per_square_foot = 165
		elif 70001 <= total_sqft <= 75000:
			cost_per_square_foot = 163
		elif 75001 <= total_sqft <= 80000:
			cost_per_square_foot = 161
		elif total_sqft >= 80001:
			cost_per_square_foot = 160
	return cost_per_square_foot

def barrier_cost(bar):
	total_cost = 0
	for hgt, total_sqft in bar.items():
		total_cost += total_sqft * cost_per_square_foot(hgt, total_sqft)
	return total_cost

def barrier_summary(barrier_analysis, barriers):
	"""Return length and total square foot of barrier or barrier system"""
	return barrier_analysis.total_length(barriers), barrier_analysis.total_square_footage(barriers)


if __name__ == '__main__':
	pass
