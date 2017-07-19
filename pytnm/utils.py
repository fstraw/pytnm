"""

Convert excel to list or tuple

"""

import shapefile
import openpyxl

wbname = '../files/tnm_inputs.xlsx'

# def tnm_feature_to_list(wbname, ws):
#     """ Returns filtered list of specified feature. Requires spreadsheet with named worksheet"""
#     wb = openpyxl.load_workbook(wbname)
#     feat_ws = wb[ws]
#     feats = []
#     for row in feat_ws.rows:
#         vals = [cell.value for cell in row]
#         feats.append(vals)
#     feats_no_header = feats[15:] #clip TNM2.5 header
#     features = []
#     for feat in feats_no_header:
#         rd_name, rd_width = rd[1], rd[2]
#         x, y, z = rd[6], rd[7], rd[8]
#         if rd_name:
#             roads.append([rd_name, rd_width, [[x, y, z]]])
#             found_road = True
#         else:
#             found_road = False
#             if not found_road:
#                 roads[-1][2].append([x, y, z])
#     return roads

def rds_to_list(wbname, ws='Roads'):
    """
    Returns filtered list of roads
    :param wb: TNM excel template (.xlsx)
    :param ws: 
    :return: 
    """
    wb = openpyxl.load_workbook(wbname)
    roads_ws = wb[ws]
    rds = []
    for row in roads_ws.rows:
        vals = [cell.value for cell in row]
        rds.append(vals)
    rds_no_header = rds[15:] #clip TNM2.5 header
    roads = []
    for rd in rds_no_header:
        rd_name, rd_width = rd[1], rd[2]
        x, y, z = rd[6], rd[7], rd[8]
        if rd_name:
            roads.append([rd_name, rd_width, [[x, y, z]]])
            found_road = True
        else:
            found_road = False
            if not found_road:
                roads[-1][2].append([x, y, z])
    return roads

def barriers_to_list(wbname, ws='Barriers'):
    """
    Returns filtered list of barriers
    :param wb: TNM excel template (.xlsx)
    :param ws: 
    :return: 
    """
    wb = openpyxl.load_workbook(wbname)
    barriers_ws = wb[ws]
    bars = []
    for row in barriers_ws.rows:
        vals = [cell.value for cell in row]
        bars.append(vals)
    bars_no_header = bars[15:] #clip TNM2.5 header
    barriers = []
    for bar in bars_no_header:
        bar_name, bar_max_height, square_foot_cost = bar[1], bar[4], bar[5]
        x, y, z = bar[13], bar[14], bar[15]
        if bar_name:
            barriers.append([bar_name, bar_max_height, square_foot_cost, [[x, y, z]]])
            found_bar = True
        else:
            found_bar = False
            if not found_bar:
                barriers[-1][3].append([x, y, z])
    return barriers

def rds_list_to_shape(rdslist, outputshp, traf_dict=None):
    w = shapefile.Writer(shapefile.POLYLINEZ)
    w.field('Rd_Name', 'C', size=32)
    w.field('Width', 'N')
    w.field('Auto', 'N')
    w.field('Medium', 'N')
    w.field('Heavy', 'N')
    w.field('MedPct', 'I')
    w.field('HvyPct', 'I')
    w.field('Speed', 'C')
    for rd in rdslist:
        w.line(parts=[rd[2]], shapeType=13)
        if not traf_dict:
            w.record(rd[0], rd[1], 0, 0, 0, 0, 0, 0)
        else:
            auto = traffic_dict[rd[0]][0]
            medium = traffic_dict[rd[0]][1]
            heavy = traffic_dict[rd[0]][2]
            medpct = traffic_dict[rd[0]][3]
            hvypct = traffic_dict[rd[0]][4]
            speed = traffic_dict[rd[0]][5]
            w.record(rd[0], rd[1], auto, medium, heavy, medpct, hvypct, speed)
    w.save(outputshp)
    return outputshp

def bars_list_to_shape(barslist, outputshp):
    w = shapefile.Writer(shapefile.POLYLINEZ)
    w.field('Bar_Name', 'C', size=32)
    w.field('MaxHeight', 'N')
    w.field('CostPerSF', 'N')
    for bar in barslist:
        w.line(parts=[bar[3]], shapeType=13)
        w.record(bar[0], bar[1], bar[2])
    w.save(outputshp)
    return outputshp

def append_tnm_traffic(wbname, ws='Traffic'):
    wb = openpyxl.load_workbook(wbname)
    traffic_ws = wb[ws]
    traf = []
    for row in traffic_ws.rows:
        vals = [cell.value for cell in row]
        traf.append(vals)
    traf_no_header = traf[14:] #clip TNM2.5 header
    traffic_vols = []
    for traffic in traf_no_header:
        rd_name, auto, medium, heavy, speed = traffic[1], traffic[5], traffic[7], traffic[9], traffic[6]
        medpct = 0
        hvypct = 0
        if rd_name:
            if medium:
                medpct = round(float(medium) / (auto + medium + heavy) * 100, 1)
            if heavy:
                hvypct = round(float(heavy) / (auto + medium + heavy) * 100, 1)
            traffic_vols.append([rd_name, auto, medium, heavy, medpct, hvypct, speed])
    return {t[0]: (t[1], t[2], t[3], t[4], t[5], t[6]) for t in traffic_vols}


if __name__ == '__main__':
    rdshp = r'../files/roads'
    barshp = r'../files/barriers'

    barslist = barriers_to_list(wbname, ws='Barriers')
    bars_list_to_shape(barslist, barshp)
    traffic_dict = append_tnm_traffic(wbname, ws='Traffic')
    rdslist = rds_to_list(wbname, ws='Roads')
    rds_list_to_shape(rdslist, rdshp, traffic_dict)
    for rd in traffic_dict:
        print rd
