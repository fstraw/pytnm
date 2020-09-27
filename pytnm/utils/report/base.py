import itertools
import openpyxl
import shapefile


class Analysis(object):
    pass

class Report(object):
    """ Reports receiver results based on shapefile of known schema
    """       

    def __init__(self, receivers, common_noise_environment=None):
        self.shp = receivers
        self.common_noise_env = common_noise_environment
        self.SENSITIVE_QUERY = ('A', 'B', 'C', 'D', 'E')   

    def _get_field_index(self, fields):
        flds = [fld for fld in fields if fld[0] != 'DeletionFlag'] # removed hidden field for proper indexing
        field_map = {}
        for index, field in enumerate(flds):
            field_name = field[0]
            field_map[field_name] = index    
        return field_map

    def _field_map(self):
        return self._get_field_index(shapefile.Reader(self.shp).fields)

    def all_receivers(self):
        with shapefile.Reader(self.shp) as sreader:
            if self.common_noise_env:
                cne = self._field_map()['assoc_bar'] 
                return [record for record in sreader.records() if record[cne] == self.common_noise_env]
            else:    
                return [record for record in sreader.records()]

    def nac_b_receivers(self):
        nac_cat = self._field_map()['nac_cat']                       
        return [record for record in self.all_receivers() if record[nac_cat] == 'B']

    def nac_c_receivers(self):
        nac_cat = self._field_map()['nac_cat']                       
        return [record for record in self.all_receivers() if record[nac_cat] == 'C']

    def nac_d_receivers(self):
        nac_cat = self._field_map()['nac_cat']                       
        return [record for record in self.all_receivers() if record[nac_cat] == 'D']

    def nac_e_receivers(self):
        nac_cat = self._field_map()['nac_cat']                       
        return [record for record in self.all_receivers() if record[nac_cat] == 'E']

    def nac_b_receptors_total(self):
        du = self._field_map()['du']
        nac_cat = self._field_map()['nac_cat']                       
        return sum([row[du] for row in self.all_receivers() if row[nac_cat] == 'B'])

    def nac_c_receptors_total(self):
        du = self._field_map()['du']
        nac_cat = self._field_map()['nac_cat']                       
        return sum([row[du] for row in self.all_receivers() if row[nac_cat] == 'C'])

    def nac_d_receptors_total(self):
        du = self._field_map()['du']
        nac_cat = self._field_map()['nac_cat']                       
        return sum([row[du] for row in self.all_receivers() if row[nac_cat] == 'D'])

    def nac_e_receptors_total(self):
        du = self._field_map()['du']
        nac_cat = self._field_map()['nac_cat']                       
        return sum([row[du] for row in self.all_receivers() if row[nac_cat] == 'E'])

    def sensitive_receivers(self):
        nac_cat = self._field_map()['nac_cat']                       
        return [record for record in self.all_receivers() if record[nac_cat] in self.SENSITIVE_QUERY]

    def sensitive_receptors_total(self):
        du = self._field_map()['du']
        nac_cat = self._field_map()['nac_cat']        
        return sum([row[du] for row in self.all_receivers() if row[nac_cat] in self.SENSITIVE_QUERY])

    def benefited_receptors(self):
        du = self._field_map()['du']
        bar_reduct = self._field_map()['bar_reduct']
        return sum([record[du] for record in self.sensitive_receivers() if record[bar_reduct] >= 5])

    def receptors_total(self):
        du = self._field_map()['du']
        return sum([row[du] for row in self.all_receivers()])

    def impacted_receivers(self):
        impacts = list(
            itertools.chain(
            self.build_b_nac_impacts(),
            self.build_c_nac_impacts(),
            self.build_d_nac_impacts(),
            self.build_e_nac_impacts(),
            self.build_b_substantial_impacts(),
            self.build_c_substantial_impacts(),
            self.build_d_substantial_impacts(),
            self.build_e_substantial_impacts()
            )
        )            
        return impacts

    def impacted_total(self):
        du = self._field_map()['du']
        return sum([impact[du] for impact in self.impacted_receivers()])

    def build_b_impacted_total(self):
        nac_cat = self._field_map()['nac_cat']
        du = self._field_map()['du']
        return sum([impact[du] for impact in self.impacted_receivers() if impact[nac_cat] == 'B'])

    def build_c_impacted_total(self):
        nac_cat = self._field_map()['nac_cat']
        du = self._field_map()['du']
        return sum([impact[du] for impact in self.impacted_receivers() if impact[nac_cat] == 'C'])

    def build_d_impacted_total(self):
        nac_cat = self._field_map()['nac_cat']
        du = self._field_map()['du']
        return sum([impact[du] for impact in self.impacted_receivers() if impact[nac_cat] == 'D'])

    def build_e_impacted_total(self):
        nac_cat = self._field_map()['nac_cat']
        du = self._field_map()['du']
        return sum([impact[du] for impact in self.impacted_receivers() if impact[nac_cat] == 'E'])

    def build_b_nac_impacts(self):        
        nac_cat = self._field_map()['nac_cat']
        bld_snd = self._field_map()['bld_snd']
        return [row for row in self.sensitive_receivers() if (row[nac_cat] == 'B' and row[bld_snd] >= 66)]

    def build_c_nac_impacts(self):        
        nac_cat = self._field_map()['nac_cat']
        bld_snd = self._field_map()['bld_snd']
        return [row for row in self.sensitive_receivers() if (row[nac_cat] == 'C' and row[bld_snd] >= 66)]

    def build_d_nac_impacts(self):        
        nac_cat = self._field_map()['nac_cat']
        bld_snd = self._field_map()['bld_snd']
        return [row for row in self.sensitive_receivers() if (row[nac_cat] == 'D' and row[bld_snd] >= 51)]

    def build_e_nac_impacts(self):        
        nac_cat = self._field_map()['nac_cat']
        bld_snd = self._field_map()['bld_snd']
        return [row for row in self.sensitive_receivers() if (row[nac_cat] == 'E' and row[bld_snd] >= 71)]

    def build_b_substantial_impacts(self):        
        nac_cat = self._field_map()['nac_cat']
        bld_snd = self._field_map()['bld_snd']
        ex_snd = self._field_map()['ex_snd']
        return [row for row in self.sensitive_receivers() if (row[nac_cat] == 'B' and (row[bld_snd] - row[ex_snd] >= 15))]

    def build_c_substantial_impacts(self):        
        nac_cat = self._field_map()['nac_cat']
        bld_snd = self._field_map()['bld_snd']
        ex_snd = self._field_map()['ex_snd']
        return [row for row in self.sensitive_receivers() if (row[nac_cat] == 'C' and (row[bld_snd] - row[ex_snd] >= 15))]

    def build_d_substantial_impacts(self):        
        nac_cat = self._field_map()['nac_cat']
        bld_snd = self._field_map()['bld_snd']
        ex_snd = self._field_map()['ex_snd']
        return [row for row in self.sensitive_receivers() if (row[nac_cat] == 'D' and (row[bld_snd] - row[ex_snd] >= 15))]

    def build_e_substantial_impacts(self):  
        nac_cat = self._field_map()['nac_cat']
        bld_snd = self._field_map()['bld_snd']
        ex_snd = self._field_map()['ex_snd']
        return [row for row in self.sensitive_receivers() if (row[nac_cat] == 'E' and (row[bld_snd] - row[ex_snd] >= 15))]

    def existing_minimum(self):
        ex_snd = self._field_map()['ex_snd']
        return min([row[ex_snd] for row in self.sensitive_receivers()])

    def existing_maximum(self):
        ex_snd = self._field_map()['ex_snd']
        return max([row[ex_snd] for row in self.sensitive_receivers()])

    def nobld_minimum(self):
        nobld_snd = self._field_map()['nobld_snd']
        return min([row[nobld_snd] for row in self.sensitive_receivers()])

    def nobld_maximum(self):
        nobld_snd = self._field_map()['nobld_snd']
        return max([row[nobld_snd] for row in self.sensitive_receivers()])

    def bld_minimum(self):
        bld_snd = self._field_map()['bld_snd']
        return min([row[bld_snd] for row in self.sensitive_receivers()])

    def bld_maximum(self):
        bld_snd = self._field_map()['bld_snd']
        return max([row[bld_snd] for row in self.sensitive_receivers()])
    
    def bld_average_change(self):
        ex_snd = self._field_map()['ex_snd']
        bld_snd = self._field_map()['bld_snd']
        diff_list = [row[bld_snd] - row[ex_snd] for row in self.sensitive_receivers()]
        return round(sum(diff_list) / len(diff_list), 1)

    def nobld_average_change(self):
        ex_snd = self._field_map()['ex_snd']
        nobld_snd = self._field_map()['nobld_snd']
        diff_list = [row[nobld_snd] - row[ex_snd] for row in self.sensitive_receivers()]
        return round(sum(diff_list) / len(diff_list), 1)

    def summary(self):
        narrative = f"A total of {self.sensitive_receptors_total()} noise sensitive receptors were analyzed. "
        narrative += f"Existing noise levels range from {self.existing_minimum()} to {self.existing_maximum()} dB(A) at {self.sensitive_receptors_total()} receptors. "
        narrative += f"No Build noise levels would range from {self.nobld_minimum()} to {self.nobld_maximum()} dB(A). "
        narrative += f"Build noise levels would range from {self.bld_minimum()} to {self.bld_maximum()} dB(A). "        
        narrative += f"Noise is predicted to change by an average of {self.nobld_average_change()} dB(A) under the No Build Alternative. "
        narrative += f"Noise is predicted to change by an average of {self.bld_average_change()} dB(A) under the Build Alternative. "
        if self.impacted_total() == 0:
            narrative += f"No receptors would be impacted under the Build Alternative."
        else:
            narrative += f"A total of {self.impacted_total()} receptors would be impacted under the Build Alternative; therefore, noise abatement was considered."
        return narrative


def create_florida_barrier_summary(xlsx, bar_length, cost_per_sq_ft):
    """ Generate Florida barrier summary based on Sound Results table for *each* barrier
    analysis in TNM Run. 
    
    Arguments:
        xlsx {String} -- Path to xlsx
    """
    xlsx = openpyxl.load_workbook(xlsx)
    barrier_summaries = xlsx.worksheets
    empty_row = [None] * 14 # empty cells, TNM standard output
    increments = [8, 10, 12, 14, 16, 18, 20, 22] #feet
    barrier_summary = []
    for idx, bar_hgt in enumerate(increments):
        table_rows = [row for row in barrier_summaries[idx].rows][19:]
        results = []
        for row in table_rows:
            cells = [cell.value for cell in row]
            if cells != empty_row:
                results.append(cells)
            else:
                break
        impacted_receptors = [row for row in results if row[9] != ' ----']
        impacted_receptors_count = sum([row[3] for row in impacted_receptors])
        min_benefit = sum([row[3] for row in impacted_receptors if 5 <= row[11] <= 5.9])
        mid_benefit = sum([row[3] for row in impacted_receptors if 6 <= row[11] <= 6.9])
        max_benefit = sum([row[3] for row in impacted_receptors if 7 <= row[11]])
        benefited_receptors = [row for row in results if row[11] >= 5]
        benefited_receptor_count = sum([row[3] for row in benefited_receptors])
        impacted_benefited_receptor_count = sum([row[3] for row in impacted_receptors if row[11] >= 5])
        impacted_not_benefited_receptor_count = benefited_receptor_count - impacted_benefited_receptor_count
        estimated_cost = round(bar_length * bar_hgt * cost_per_sq_ft)
        try:
            average_benefit_reduction = round(sum([row[11] for row in benefited_receptors]) / len(benefited_receptors), 1)
            cost_per_benefit = round(estimated_cost / benefited_receptor_count)
        except ZeroDivisionError:
            average_benefit_reduction = 0        
            cost_per_benefit = 0
        barrier_summary.append((
            bar_hgt, '{:,.0f}'.format(bar_length), impacted_receptors_count, min_benefit, mid_benefit, max_benefit, 
            impacted_benefited_receptor_count, impacted_not_benefited_receptor_count, benefited_receptor_count, 
            average_benefit_reduction, estimated_cost, cost_per_benefit))
    return barrier_summary

def summarize_cne(receivers, cne):
    r = Report(receivers, cne)    
    return (
        cne, 
        'LOCATION', 
        r.nac_b_receptors_total(), 
        r.nac_c_receptors_total(),
        r.nac_e_receptors_total(),
        r.existing_minimum(),
        r.existing_maximum(),
        r.nobld_minimum(),
        r.nobld_maximum(),
        r.bld_minimum(),
        r.bld_maximum(),
        r.build_b_impacted_total(),
        r.build_c_impacted_total(),
        r.build_e_impacted_total(),
        'WARRANTED'
    )

def create_cne_summary_table(receivers, cne_list, xlsx):
    wb = openpyxl.Workbook()
    ws = wb.active
    hdrs = ('CNE', 'Location', 'NAC B', 'NAC C', 'NAC E', 'Min', 'Max', 'Min', 'Max', 'Min', 'Max', 'NAC B', 'NAC C', 'NAC E', 'Abatement Warranted')
    ws.append(hdrs)
    for cne in cne_list:
        cne_summary = summarize_cne(receivers, cne)
        ws.append(cne_summary)    
    wb.save(filename=xlsx)

def create_barrier_summary_table(barrier_summary, xlsx):
    """Generate .xlsx file of barrier results
    
    Arguments:
        barrier_summary {}
        xlsx {String} -- Full output path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    hdrs = ("Barrier Height", "Barrier Length", "Impacted Receptors", 
    "5 - 5.9 dB(A)", "6 - 6.9 dB(A)", "7+", "Impacted", "Not Impacted", 
    "Total", "Average Benefited Reduction", "Estimated Cost", "Cost per Benefit", "Feasible and Reasonable")
    ws.append(hdrs)
    for barrier_design in barrier_summary:
        ws.append(barrier_design)
    wb.save(filename=xlsx)