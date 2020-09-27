"""

Convert excel to list or tuple

"""
import os
import shapefile
import openpyxl


def bar_point_id_xy_dict(wbname, ws='Barriers'):
    wb = openpyxl.load_workbook(wbname)
    barriers_ws = wb[ws]
    bars = []
    for row in barriers_ws.rows:
        vals = [cell.value for cell in row]
        bars.append(vals)
    bars_no_header = bars[15:] #clip TNM2.5 header
    return {row[11]: (row[13], row[14], row[15]) for row in bars_no_header}

def get_last_barrier_seg_pointid(wbname, ws='Barriers'):
    """
    Returns last point for each barrier - temporary hack
    :param wbname:
    :param ws:
    :return:
    """
    wb = openpyxl.load_workbook(wbname)
    barriers_ws = wb[ws]
    bars = []
    lastpointdict = {}
    for row in barriers_ws.rows:
        vals = [cell.value for cell in row]
        bars.append(vals)
    bars_no_header = bars[15:] #clip TNM2.5 header
    for row in bars_no_header:
        barname = row[1]
        lastindicator = row[17]
        if barname:
            currentbar = barname
        if lastindicator == None:
            lastpointdict[currentbar] = row[11]
    return lastpointdict

def rds_to_list(wbname, ws='Roads'):
    """
    Returns filtered list of roads
    :param wb: TNM excel template (.xlsx)
    :param ws: 
    :return: 
    """
    wb = openpyxl.load_workbook(wbname)
    roads_ws = wb[ws]
    rds = []
    for row in roads_ws.rows:
        vals = [cell.value for cell in row]
        rds.append(vals)
    rds_no_header = rds[15:] #clip TNM2.5 header
    roads = []
    for rd in rds_no_header:
        rd_name, rd_width = rd[1], rd[2]
        x, y, z = rd[6], rd[7], rd[8]
        #check for empty end rows
        if None in (x, y, z):
            break
        if rd_name:
            roads.append([rd_name, rd_width, [[x, y, z]]])
            found_road = True
        else:
            found_road = False
            if not found_road:
                roads[-1][2].append([x, y, z])
    return roads

def barriers_to_list(wbname, ws='Barriers'):
    """
    Returns filtered list of barriers
    :param wb: TNM excel template (.xlsx)
    :param ws: 
    :return: 
    """
    wb = openpyxl.load_workbook(wbname)
    barriers_ws = wb[ws]
    bars = []
    for row in barriers_ws.rows:
        vals = [cell.value for cell in row]
        bars.append(vals)
    bars_no_header = bars[15:] #clip TNM2.5 header
    barriers = []
    for bar in bars_no_header:
        bar_name, bar_max_height, square_foot_cost = bar[1], bar[4], bar[5]
        x, y, z = bar[13], bar[14], bar[15]
        #check for empty end rows
        if None in (x, y, z):
            break
        if bar_name:
            barriers.append([bar_name, bar_max_height, square_foot_cost, [[x, y, z]]])
            found_bar = True
        else:
            found_bar = False
            if not found_bar:
                barriers[-1][3].append([x, y, z])
    return barriers

def group_barrier_segments_by_height(wbname, ws='Barrier_Segments'):
    """
    Parses barrier into individual segment lengths, based on Barrier Segments output (TNM)
    :param wbname:
    :param ws:
    :return:
    """
    wb = openpyxl.load_workbook(wbname)
    barriers_ws = wb[ws]
    bars = []
    for row in barriers_ws.rows:
        vals = [cell.value for cell in row]
        bars.append(vals)
    bars_no_header = bars[15:] #clip TNM2.5 header
    barriers = []
    #coupled - consider revisiting approach
    XYZ_DICT = bar_point_id_xy_dict(wbname)
    LAST_POINTS_DICT = get_last_barrier_seg_pointid(wbname) #consider revising approach
    i = 0
    while i < len(bars_no_header): #avoid index error
        bar_name, bar_height = bars_no_header[i][1], bars_no_header[i][7]
        try:
            pointid1 = bars_no_header[i][4]
            x1, y1, z1 = XYZ_DICT[pointid1]  # should return tuple
            pointid2 = bars_no_header[i+1][4]
            x2, y2, z2 = XYZ_DICT[pointid2]
        except IndexError: #account for end of list
            pointid2 = LAST_POINTS_DICT[current_bar] # should be assigned by the time this error throws
            x2, y2, z2 = XYZ_DICT[pointid2]
            break
        except KeyError: #check for empty row
            bar_height = bars_no_header[i-1][7]
            pointid1 = bars_no_header[i-1][4]
            x1, y1, z1 = XYZ_DICT[pointid1]  # should return tuple
            pointid2 = LAST_POINTS_DICT[current_bar]  # should be assigned by the time this error throws
            x2, y2, z2 = XYZ_DICT[pointid2]
        if not pointid1:
            bar_height = bars_no_header[i-1][7]
            pointid1 = bars_no_header[i-1][4]
            pointid2 = LAST_POINTS_DICT[current_bar]
            x1, y1, z1 = XYZ_DICT[pointid1]
            x2, y2, z2 = XYZ_DICT[pointid2]
        if bar_name:
            current_bar = bar_name
        if bar_height == 0:
            i+=1
            continue
        barriers.append([current_bar, pointid1, bar_height, [[x1, y1, z1], [x2, y2, z2]]])
        i+=1
    return barriers

def rds_list_to_shape(rdslist, outputshp, traf_dict=None):
    w = shapefile.Writer(shapeType=shapefile.POLYLINEZ)
    w.field('Rd_Name', 'C', size=32)
    w.field('Width', 'N')
    w.field('Auto', 'N')
    w.field('Medium', 'N')
    w.field('Heavy', 'N')
    w.field('MedPct', 'N')
    w.field('HvyPct', 'N')
    w.field('Speed', 'C')
    for rd in rdslist:
        w.line(parts=[rd[2]], shapeType=13)
        if not traf_dict:
            w.record(rd[0], int(rd[1]), 0, 0, 0, 0, 0, 0)
        else:
            auto = traf_dict[rd[0]][0]
            medium = traf_dict[rd[0]][1]
            heavy = traf_dict[rd[0]][2]
            medpct = traf_dict[rd[0]][3]
            hvypct = traf_dict[rd[0]][4]
            speed = traf_dict[rd[0]][5]
            w.record(rd[0], int(rd[1]), auto, medium, heavy, medpct, hvypct, speed)
    w.save(outputshp)
    return outputshp

def bars_list_to_shape(barslist, outputshp):
    w = shapefile.Writer(shapefile.POLYLINEZ)
    w.field('Bar_Name', 'C', size=32)
    w.field('MaxHeight', 'N')
    w.field('CostPerSF', 'N')
    for bar in barslist:
        w.line(parts=[bar[3]], shapeType=13)
        w.record(bar[0], int(bar[1]), bar[2])
    w.save(outputshp)
    return outputshp

def bar_segs_list_to_shape(barslist, outputshp):
    w = shapefile.Writer(shapefile.POLYLINEZ)
    w.field('Bar_Name', 'C', size=32)
    w.field('PointID', 'C', size=10)
    w.field('AVGHEIGHT', 'N')
    for bar in barslist:
        w.line(parts=[bar[3]], shapeType=13)
        w.record(bar[0], bar[1], int(bar[2]))
    w.save(outputshp)
    return outputshp

def append_tnm_traffic(wbname, ws='Traffic'):
    wb = openpyxl.load_workbook(wbname)
    traffic_ws = wb[ws]
    traf = []
    for row in traffic_ws.rows:
        vals = [cell.value for cell in row]
        traf.append(vals)
    traf_no_header = traf[14:] #clip TNM2.5 header
    traffic_vols = []
    for traffic in traf_no_header:
        rd_name, auto, medium, heavy, speed = traffic[1], traffic[5], traffic[7], traffic[9], traffic[6]
        medpct = 0
        hvypct = 0
        if rd_name:
            if medium:
                medpct = round(float(medium) / (auto + medium + heavy) * 100, 1)
            if heavy:
                hvypct = round(float(heavy) / (auto + medium + heavy) * 100, 1)
            traffic_vols.append([rd_name, auto, medium, heavy, medpct, hvypct, speed])            
    return {t[0]: (t[1], t[2], t[3], t[4], t[5], t[6]) for t in traffic_vols}


if __name__ == '__main__':
    pass